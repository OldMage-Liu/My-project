#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
抖音评论提取工具（完整版本）
- 保留所有原始定位器
- 解决特殊字符/空用户名导致的对齐问题
- 支持Excel导出并优化格式
"""
from math import nan
import time
import pandas as pd
import os as os_alias
import random as rnd
import unicodedata
from datetime import datetime
from openpyxl.styles import Alignment
from playwright.sync_api import sync_playwright as playwright_alias

# 路径配置
USER_PROFILE_PATH = os_alias.path.join(os_alias.getcwd(), "chrome_profile")
EXCEL_SAVE_DIR = os_alias.path.join(os_alias.getcwd(), "抖音评论_Excel数据")
os_alias.makedirs(EXCEL_SAVE_DIR, exist_ok=True)

# 浏览器启动参数
CHROME_LAUNCH_ARGS = [
    "--disable-blink-features=AutomationControlled",
    "--disable-web-security",
    "--disable-site-isolation-trials",
    "--disable-features=IsolateOrigins,site-per-process",
    "--disable-infobars",
    "--window-size=1366,768",
    "--no-sandbox",
    "--disable-dev-shm-usage",
    "--disable-gpu"
]


def _norm_text(s: str) -> str:
    """清洗文本：处理特殊字符、不可见字符和空值"""
    if s is None:
        return ""
    try:
        s = unicodedata.normalize("NFKC", s)  # 标准化Unicode字符
    except Exception:
        pass
    # 移除零宽空格、软连字符等不可见字符
    invisible_chars = ["\u200b", "\u200c", "\u200d", "\ufeff", "\r", "\n"]
    for c in invisible_chars:
        s = s.replace(c, "")
    return s.strip()


def is_xpath(selector: str) -> bool:
    """判断选择器是否为XPath"""
    return selector.startswith(('//', '/html', 'xpath='))


def get_all_elements_sync(page, sel1, sel2):
    """合并两个选择器的元素列表（保留原始逻辑）"""
    loc1 = page.locator(f'xpath={sel1}' if is_xpath(sel1) else sel1)
    loc2 = page.locator(f'xpath={sel2}' if is_xpath(sel2) else sel2)
    return loc1.element_handles() + loc2.element_handles()


def save_to_excel(merged_data, video_num, video_title=""):
    """保存数据到Excel并优化格式"""
    excel_data = []
    for idx, (name, ip, comment) in enumerate(merged_data, 1):
        # 统一处理空值显示
        clean_name = _norm_text(name) if (name not in [nan, None, ""]) else f"匿名用户_{idx}"
        clean_ip = _norm_text(ip).replace("IP属地：", "") if (ip not in [nan, None, ""]) else "未知"
        clean_comment = _norm_text(comment) if (comment not in [nan, None, ""]) else "无内容"

        excel_data.append({
            "序号": idx,
            "用户名": clean_name,
            "IP属地": clean_ip,
            "评论内容": clean_comment
        })

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os_alias.path.join(EXCEL_SAVE_DIR, f"视频{video_num}_评论_{timestamp}.xlsx")

    # 保存并设置Excel格式
    df = pd.DataFrame(excel_data)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=f"视频{video_num}", index=False)
        ws = writer.sheets[f"视频{video_num}"]

        # 列宽设置
        ws.column_dimensions["A"].width = 8  # 序号
        ws.column_dimensions["B"].width = 25  # 用户名
        ws.column_dimensions["C"].width = 15  # IP属地
        ws.column_dimensions["D"].width = 60  # 评论内容

        # 自动换行和对齐方式
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                    horizontal="left"
                )

        # 添加视频标题
        if video_title:
            ws.insert_rows(1)
            ws["A1"] = f"视频标题：{_norm_text(video_title)}"
            ws.merge_cells("A1:D1")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    print(f"✅ 数据已保存至：{excel_path}")
    return excel_path


def run_main():
    file_counter = 0  # 视频计数器
    with playwright_alias() as pw:
        # 启动浏览器（持久化上下文）
        browser_context = pw.chromium.launch_persistent_context(
            user_data_dir=USER_PROFILE_PATH,
            channel="chrome",
            headless=False,
            args=CHROME_LAUNCH_ARGS,
            locale="zh-CN",
            timezone_id="Asia/Shanghai",
            ignore_default_args=["--enable-automation"]
        )

        page = browser_context.new_page()
        # 规避反爬检测
        page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            window.chrome = window.chrome || {};
        """)

        # 1. 打开抖音并搜索
        page.goto("https://www.douyin.com/", timeout=10000)
        page.wait_for_timeout(1000)  # 等待首页加载
        # 定位搜索框（原始定位器）
        search_input = page.locator(
            '//*[@id="douyin-header"]/div[1]/header/div/div/div[1]/div/div[2]/div/div[1]/input')
        search_input.fill('手机推荐')  # 搜索关键词
        # 点击搜索按钮（原始定位器）
        page.locator('//*[@id="douyin-header"]/div[1]/header/div/div/div[1]/div/div[2]/div/button/span').click()
        page.wait_for_timeout(2000)  # 等待搜索结果加载
        # input()

        # 2. 点击第一个视频（原始定位器，强制选择第一个）
        video_locator = page.locator(
            '#search-content-area > div > div.LI8kV7Vf > div.oo_LwYT3.NpAMJIe0.p576IkAN > div.NqYoxcpn > div > div > span:nth-child(3)'
        )
        video_locator.first.click()  # 避免多元素匹配错误
        page.wait_for_timeout(2000)  # 等待视频加载

        # 3. 等待右侧视频列表加载（原始定位器）
        right_panel = page.locator('//*[@id="douyin-right-container"]/div[3]')
        right_panel.wait_for(state='visible', timeout=10000)

        # 4. 滚动加载更多视频（右侧列表）
        for _ in range(30):
            for __ in range(5):
                right_panel.evaluate('el => el.scrollTop += el.clientHeight / 5')
                page.wait_for_timeout(rnd.randint(50, 100))  # 随机延迟模拟人工
            page.wait_for_timeout(300)
        page.wait_for_timeout(1000)

        # 5. 获取视频列表元素（原始定位器）
        video_cards = right_panel.locator('//*[@id="search-result-container"]/div[2]/ul/li/div').all()
        video_titles = right_panel.locator(
            '//*[@id="search-result-container"]/div[2]/ul/li/div/a/div/div[2]/div/div[1]').all()

        # 6. 逐视频处理（限制数量，避免过载）
        max_video_count = 2  # 可修改处理的视频数量
        for card, title in zip(video_cards[:max_video_count], video_titles[:max_video_count]):
            file_counter += 1
            video_title = _norm_text(title.text_content())  # 视频标题
            print(f"\n===== 开始处理视频 {file_counter}：{video_title[:20]} =====")

            # 6.1 点击视频（确保可见）
            card.scroll_into_view_if_needed()
            card.click()
            page.wait_for_timeout(2000)  # 等待视频页加载
            page.keyboard.press("x")  # 关闭弹幕

            # 6.2 等待评论区加载（原始定位器）
            comment_container = page.locator(
                '#merge-all-comment-container > div > div.Rwb9ssMc.comment-mainContent.ufktjxUm')
            comment_container.wait_for(state='visible', timeout=10000)

            # 6.3 滚动加载评论（获取更多数据）
            for _ in range(30):
                for __ in range(5):
                    comment_container.evaluate('el => el.scrollTop += el.clientHeight / 5')
                    page.wait_for_timeout(rnd.randint(50, 100))
                page.wait_for_timeout(500)  # 等待新评论加载

            # 6.4 展开回复（原始定位器）
            expand_buttons = page.query_selector_all(
                '#merge-all-comment-container > div > div.Rwb9ssMc.comment-mainContent.ufktjxUm > div > div > div.EpsntdUI > button')
            for btn in expand_buttons:
                try:
                    btn.click()
                    page.wait_for_timeout(300)  # 等待回复展开
                except Exception as e:
                    print(f"展开回复失败：{e}")
                    continue

            # 6.5 滚动到评论顶部（重新梳理顺序）
            current_top = comment_container.evaluate('el => el.scrollTop')
            while current_top > 0:
                comment_container.evaluate('el => el.scrollTop -= 200')  # 逐步上滚
                page.wait_for_timeout(300)
                current_top = comment_container.evaluate('el => el.scrollTop')
            page.wait_for_timeout(2000)  # 稳定页面

            # 7. 提取评论、用户名、用户链接（原始定位器）
            # 评论定位器
            comment_span_sel = '//*[@id="merge-all-comment-container"]/div/div[3]/div/div/div[2]/div/div[2]/span'
            comment_span_sel1 = '#merge-all-comment-container > div > div.Rwb9ssMc.comment-mainContent.ufktjxUm > div > div > div.EpsntdUI > div.CCbmLKh0.replyContainer > div > div.EpsntdUI > div > div.C7LroK_h > span'
            comment_elements = get_all_elements_sync(page, comment_span_sel, comment_span_sel1)

            # 用户名定位器
            username_span_sel = '//*[@id="merge-all-comment-container"]/div/div[3]/div/div/div[2]/div/div[1]/div[1]/div/a/span/span/span/span/span/span/span'
            username_span_sel1 = '//*[@id="merge-all-comment-container"]/div/div[3]/div/div/div[2]/div[2]/div/div/div/div/div/div/a[1]/span/span/span/span/span/span/span'
            username_elements = get_all_elements_sync(page, username_span_sel, username_span_sel1)

            # 用户链接定位器
            user_link_sel1 = '//*[@id="merge-all-comment-container"]/div/div[3]/div/div/div[2]/div/div/div[1]/div/a'
            link_elements = page.locator(user_link_sel1).element_handles()

            # 8. 评论-用户名对齐（核心优化）
            # 8.1 提取评论文本
            comment_texts = []
            for el in comment_elements:
                if not el:
                    comment_texts.append(nan)
                    continue
                try:
                    txt = (el.evaluate("e => e.innerText") or "").strip()
                    comment_texts.append(txt if txt else nan)
                except Exception as e:
                    print(f"评论提取失败：{e}")
                    comment_texts.append(nan)

            # 8.2 提取用户名（清洗特殊字符和空值）
            username_texts = []
            for el in username_elements:
                if not el:
                    username_texts.append(nan)
                    continue
                try:
                    txt = (el.evaluate("e => e.innerText") or "").strip()
                    txt = _norm_text(txt)  # 清洗特殊字符
                    # 处理空用户名（特殊字符头像/无名称）
                    if not txt or txt.isspace() or txt in ["", " ", "\u200b"]:
                        username_texts.append(nan)
                    else:
                        username_texts.append(txt)
                except Exception as e:
                    print(f"用户名提取失败：{e}")
                    username_texts.append(nan)

            # 8.3 对齐长度（以评论数为基准，确保数量一致）
            max_len = len(comment_texts)
            if len(username_texts) > max_len:
                username_texts = username_texts[:max_len]  # 截断过长列表
            elif len(username_texts) < max_len:
                username_texts.extend([nan] * (max_len - len(username_texts)))  # 补充空值

            # 对齐检查
            print(f"对齐后 - 评论数: {len(comment_texts)}, 用户名数: {len(username_texts)}")

            # 9. 提取IP和用户名称（从用户主页）
            ips = []
            names = []
            for el in link_elements:
                try:
                    with page.context.expect_page() as new_pm:
                        el.click(timeout=5000)  # 点击用户头像/名称
                    user_page = new_pm.value
                    user_page.wait_for_timeout(1000)  # 等待用户页加载

                    # 提取IP（原始定位器）
                    ip_el = user_page.locator('//*[@id="user_detail_element"]/div/div[2]/div[2]/p')
                    ip = ip_el.text_content() if ip_el.count() > 0 else nan

                    # 提取用户主页名称（原始定位器）
                    name_el = user_page.locator(
                        '#user_detail_element > div > div.a3i9GVfe.nZryJ1oM._6lTeZcQP.y5Tqsaqg > div.IGPVd8vQ > div.HjcJQS1Z > h1 > span > span')
                    name = _norm_text(name_el.text_content()) if name_el.count() > 0 else nan

                    ips.append(ip)
                    names.append(name)
                    user_page.close()
                except Exception as e:
                    print(f"IP/名称提取失败：{e}")
                    ips.append(nan)
                    names.append(nan)

            # 10. 数据合并（多级匹配，解决错位）
            # 10.1 建立用户名→评论映射
            full_comment_map = {}
            for uname, cmt in zip(username_texts, comment_texts):
                if uname not in [nan, None, ""]:
                    clean_uname = _norm_text(str(uname))
                    if clean_uname:
                        full_comment_map[clean_uname] = cmt

            # 10.2 建立名称→IP映射
            ip_map = {}
            for n, ip in zip(names, ips):
                if n not in [nan, None, ""]:
                    clean_n = _norm_text(str(n))
                    if clean_n:
                        ip_map[clean_n] = ip

            # 10.3 多级匹配合并数据
            merged = []
            for n in names:
                if n in [nan, None, ""]:
                    merged.append((n, nan, nan))
                    continue

                # 清洗当前名称，统一格式
                clean_n = _norm_text(str(n))
                # 获取对应的IP
                ip_val = ip_map.get(clean_n, nan)
                # 初始化评论（默认未匹配）
                comment_val = nan

                # ① 直接匹配：用户名完全一致
                if clean_n in full_comment_map:
                    comment_val = full_comment_map[clean_n]
                # ② 去除后缀匹配：处理带"抖音号：XXX"等后缀的用户名
                else:
                    pure_name = clean_n.split("抖音号：")[0].split("IP属地：")[0].strip()
                    if pure_name in full_comment_map:
                        comment_val = full_comment_map[pure_name]
                    # ③ 部分包含匹配：名称存在包含关系（如"用户ABC"包含"ABC"）
                    else:
                        for key in full_comment_map.keys():
                            if key and (key in clean_n or clean_n in key):
                                comment_val = full_comment_map[key]
                                break

                # 添加到合并结果
                merged.append((clean_n, ip_val, comment_val))

            # 10.4 补充未匹配的评论（避免遗漏）
            for uname, cmt in full_comment_map.items():
                if uname and cmt not in [nan, ""]:
                    # 检查该评论是否已在结果中
                    is_duplicate = False
                    for m_name, _, m_cmt in merged:
                        if m_cmt == cmt:  # 按评论内容去重
                            is_duplicate = True
                            break
                    if not is_duplicate:
                        merged.append((uname, nan, cmt))  # IP为空但保留评论

            print(f"合并后数据总量: {len(merged)} 条")

            # 11. 保存到Excel
            save_to_excel(merged, file_counter, video_title)
            # input()
            # 12. 关闭当前视频，返回列表页
            try:
                close_btn = page.locator(
                    '#douyin-right-container > div:nth-child(4) > div > div:nth-child(1) > div.uRH5Oxnw.isDark > div')
                close_btn.wait_for(state='visible', timeout=5000)
                close_btn.click()
            except Exception as e:
                print(f"关闭视频失败，尝试按ESC：{e}")
                page.keyboard.press('esc')  # 备选方案：按ESC关闭
            page.wait_for_timeout(1000)  # 等待返回列表

        # 13. 处理完毕，关闭浏览器
        print("\n===== 所有视频处理完成 =====")
        browser_context.close()
        print(f"所有数据已保存至文件夹：{EXCEL_SAVE_DIR}")


if __name__ == "__main__":
    run_main()
